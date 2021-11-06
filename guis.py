from tkinter import *
from openpyxl import load_workbook
f = "nono.xlsx"
wb = load_workbook(f)
ws = wb.worksheets[0]
ws.title="ApnaBank"
ws.sheet_properties.tabColor = "107200"
homeWindow = Tk()
homeWindow.geometry('1000x600')
homeWindow.title('bank')

def creAcc() :
    f = "nono.xlsx"
    wb = load_workbook(f)
    ws = wb.worksheets[0]
    ws.title="ApnaBank"
    ws.sheet_properties.tabColor = "107200"
    Phno = StringVar()
    add = StringVar()
    nam = StringVar()
    pin = StringVar()
    fr1.destroy()
    fr2 = Frame(homeWindow)    
    l1 = Label(fr2, text = "Enter Phone.No :").grid(row = 0, column = 0)
    e1 = Entry(fr2, textvariable = Phno).grid(row = 0,column =1)
    l2 = Label(fr2, text = "Adress :").grid(row = 1, column = 0)
    e2 = Entry(fr2, textvariable = add).grid(row = 1,column =1)
    l3 = Label(fr2, text = "Name :").grid(row = 2, column = 0)
    e3 = Entry(fr2, textvariable = nam).grid(row = 2,column =1)
    l4 = Label(fr2, text = "Pin :").grid(row = 3, column = 0)
    e4 = Entry(fr2, textvariable = pin).grid(row = 3,column =1)
    def ca():
        a1 = nam.get()
        b1 = add.get()
        c1 = int(pin.get())
        d1 = int(Phno.get())
        for c in ws['D']:
            def x(p,o,i,u):
                accBal = 0
                q=(p,o,i,u,accBal)
                ws.append(q)
                fr2.destroy()
                fr3 = Frame(homeWindow)
                l1 = Label(fr3, text = "created").grid(row = 0, column = 0)
                fr3.pack()
                wb.save(f)
                                        
            if d1 == c.value:
                fr2.destroy()
                fr3 = Frame(homeWindow)
                l1 = Label(fr3, text = "acc already exist").grid(row = 0, column = 0)
                fr3.pack()
                break
        else:
            x(a1,b1,c1,d1)
    b1 = Button(fr2,text = "Submit" , command = ca ).grid(row =4 , column = 1)
    fr2.pack()
def logg():
    fr1.destroy() 
    fr2 = Frame(homeWindow)
    fr2.configure(background = 'blue')
    f = "nono.xlsx"
    wb = load_workbook(f)
    ws = wb.worksheets[0]
    ws.title="ApnaBank"
    ws.sheet_properties.tabColor = "107200"
    pn = StringVar()
    pp = StringVar()
    l1 = Label(fr2, text = "Enter Phone.No :").grid(row = 0, column = 0)
    e1 = Entry(fr2, textvariable = pn).grid(row = 0,column =1)
    l2 = Label(fr2, text = "Pin :").grid(row = 1, column = 0)
    e2 = Entry(fr2, textvariable = pp).grid(row = 1,column =1)
    def auth(o):
        fr2.destroy()
        fr4 = Frame(homeWindow)
        def depo():
            fr4.destroy()
            fr5 = Frame(homeWindow)
            def kitna(mo):
                mainBal = ws.cell(row=o,column=5)
                tq=ws[f"E{o}"]=mainBal.value + mo
                ll2 = Label(fr5, text = f"sucessful, bal is :-> {tq}").grid(row=5,column =1)
                wb.save(f)
            moni = StringVar()
            print(moni.get())
            l1 = Label(fr5, text = "Money to depoit :").grid(row = 0, column =0)
            e1 = Entry(fr5,textvariable = moni).grid(row = 0, column =1)
            def tata():
                kitna(int(moni.get()))
            bb1 = Button(fr5, text = "Deposit",command = tata).grid(row = 1,column =1)
            fr5.pack()
        def withdw():
            fr4.destroy()
            fr5 = Frame(homeWindow)
            def kitna(mo):
                mainBal = ws.cell(row=o,column=5)
                tq=ws[f"E{o}"]=mainBal.value - mo
                ll2 = Label(fr5, text = f"sucessful, bal is :-> {tq}").grid(row=5,column =1)
                wb.save(f)
            moni = StringVar()
            print(moni.get())
            l1 = Label(fr5, text = "Money to withdraw :").grid(row = 0, column =0)
            e1 = Entry(fr5,textvariable = moni).grid(row = 0, column =1)
            def tata():
                kitna(int(moni.get()))
            bb1 = Button(fr5, text = "Withdorh",command = tata).grid(row = 1,column =1)
            fr5.pack()
        def bala():
            fr4.destroy()
            frf =Frame(homeWindow)
            mainBal = ws.cell(row=o,column=5)
            ll2 = Label(frf,text =f"you have {mainBal.value}.Rs in your account").grid(row =0 ,column =0)
            frf.pack()    
        def edi():
            fr4.destroy()
            fre = Frame(homeWindow)
            
            def naam(o):
                fre.destroy()
                frn = Frame(homeWindow)
                namaa = StringVar()
                def chnaga():
                    new = ws.cell(row=o,column=1)
                    new.value = namaa.get()
                    wb.save(f)
                    laka = Label(frn, text = f"Name changed, new name is {new.value}").grid(row = 9,column =1)
                x=f"Your old Name is {ws.cell(row=o,column=1).value}"
                y=" Enter new Name :-> "
                ll = Label(frn, text = x).grid(row =0, column =0)
                l2= Label(frn , text = y).grid(row= 1, column = 0)
                ee =Entry(frn, textvariable = namaa).grid(row = 1, column = 1)
                def dost():
                    chnaga(namaa.get())
                  
                b0 = Button(frn , text = "Change Name",command = chnaga).grid(row = 2, column =1 )
                frn.pack()      
            def addr(r):
                fre.destroy()
                fra = Frame(homeWindow)
                addd = StringVar()
                def bhnga(w):
                    new = ws.cell(row=r,column=2)
                    new.value=w
                    wb.save(f)
                    baka = Label(fra, text = f"Address changed, new Address is {w}").grid(row = 9,column =1)
                x=f"Your old Address is {ws.cell(row=r,column=2).value}"
                y=" Enter new Address :-> "
                ll = Label(fra, text = x).grid(row =0, column =0)
                l2= Label(fra, text = y).grid(row= 1, column = 0)
                ee =Entry(fra, textvariable = addd).grid(row = 1, column = 1)
                def shtru():
                    bhnga(addd.get())
                  
                b0 = Button(fra,text = "Change Address",command = shtru).grid(row = 2, column =1 )
                fra.pack()      
            def pion(r):
                fre.destroy()
                fra = Frame(homeWindow)
                addd = StringVar()
                def bhnga(w):
                    new = ws.cell(row=r,column=3)
                    new.value=w
                    wb.save(f)
                    baka = Label(fra, text = f"Pin changed, new Pin is {w}").grid(row = 9,column =1)
                x=f"Your old Pin is {ws.cell(row=r,column=3).value}"
                y=" Enter new Pin :-> "
                ll =Label(fra, text = x).grid(row =0, column =0)
                l2= Label(fra, text = y).grid(row= 1, column = 0)
                ee =Entry(fra, textvariable = addd).grid(row = 1, column = 1)
                def shtru():
                    bhnga(int(addd.get()))
                  
                b0 = Button(fra,text = "Change Pin",command = shtru).grid(row = 2, column =1 )
                fra.pack()  
                                  
            def ediA():
                addr(o)           
            def ediN():
                naam(o)
            def ediP():
                pion(o)
            b1 = Button(fre , text = "Edit Name",command = ediN).grid(row = 1 , column = 0)
            b2 = Button(fre , text = "Edit Adress",command =ediA).grid(row = 2 , column = 0)
            b3 = Button(fre , text = "Edit pin ",command = ediP).grid(row = 3, column = 0)
            fre.pack()
        b1 = Button(fr4 , text = "Deposit",command = depo).grid(row = 0 , column = 0)
        b2 = Button(fr4 , text = "Withdraw",command = withdw).grid(row = 0 , column = 1)
        b3 = Button(fr4 , text = "Edit Details",command = edi ).grid(row = 1 , column = 0)
        b4 = Button(fr4 , text = "Check Balance",command = bala ).grid(row = 1 , column = 1)
        fr4.pack()  
    def lala():
        ph1=int(pn.get())
        pi1=int(pp.get())
        a = 1
        while True:
            n=ws.cell(row=a,column=4)
            m=ws.cell(row=a,column=3)
            if (n.value==ph1):
                auth(a)       
                break
            a+=1
            if(n.value == None):
                fr2.destroy()
                fr3 = Frame(homeWindow)
                ll1 = Label(fr3,text = "Account doesnt exist").grid(row = 1, column = 1)
                fr3.pack()
                break
    b1 = Button(fr2 , text = "submit", command = lala).grid(row = 3 , column = 1)
    fr2.pack()

l1=Label(homeWindow, text = 'Welcome to bank',font=('Helvatical bold',40)).pack()
fr1 =Frame(homeWindow)
b1=Button(fr1, text="Login",padx=25, command = logg).grid(row = 0, column=1)
b2=Button(fr1, text="Register",padx=25,command = creAcc).grid(row = 0, column =2)
fr1.pack()
homeWindow.configure(background='orange')
homeWindow.mainloop()
